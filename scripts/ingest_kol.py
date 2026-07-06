#!/usr/bin/env python3
"""Ingest a KOL rate-card workbook into the canonical ARGUS KOL directory.

The Mastersheet is a paid-influencer price book: for each KOL it records the
platform(s), audience size, niche, quality rating, and — crucially — what they
charge and what you get for it. ARGUS uses this to estimate how much a project
spent on non-organic promotion and to spot clusters of KOLs hired together.

This ingester is deliberately re-runnable. Point it at a new workbook and it
UPSERTS every row into src/data/kol/directory.json keyed by (region, name),
so uploading more sheets grows the directory instead of clobbering it.

    python3 scripts/ingest_kol.py "/path/to/KOL Mastersheet.xlsx"
    python3 scripts/ingest_kol.py sheet.xlsx --replace   # drop old rows first

Requires openpyxl (pip install openpyxl).
"""
from __future__ import annotations
import argparse, hashlib, json, re, sys, unicodedata
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src" / "data" / "kol" / "directory.json"

# Two sheet layouts appear in the wild. We detect by header, not position.
LAYOUT_STANDARD = {  # header text (upper) -> canonical field
    "KOL": "name", "PLATFORM": "platforms", "TIER": "tier",
    "FOLLOWERS": "followers", "CATEGORY": "categories", "LANGUAGE": "language",
    "RATING": "rating", "PRICE RANGE": "price_range", "PRICING": "pricing",
    "DELIVERABLES": "deliverables",
}
LAYOUT_LATAM = {  # LATAM sheet
    "KOL NAME": "name", "COUNTRY": "country", "PLATFORM": "platforms",
    "LINK": "link", "FOLLOWERS": "followers", "NICHE": "categories",
    "LANGUAGE": "language", "DELIVERABLES": "deliverables", "PRICE (USD)": "pricing",
}
LAYOUT_WEB2 = {  # WEB2 sheet
    "NICHE": "categories", "NAME": "name", "LINK": "link",
    "PLATFORMS": "platforms", "FOLLOWERS": "followers", "PRICE": "pricing",
}

PRICE_ON_REQUEST = {"POR", "POQ", "PRICE ON REQUEST", "-", ""}


def clean(v):
    if v is None:
        return ""
    if isinstance(v, str):
        # normalize unicode + collapse whitespace, strip zero-width marks
        v = unicodedata.normalize("NFKC", v).replace("‏", "").replace("‎", "")
        return v.strip()
    return v


def parse_followers(v):
    """Return (int|None, raw_str). Handles 323.3K, 2M, 51.847 (=51,847),
    983.4k, 100K+, '276 Members', 10200, '23k'."""
    raw = clean(v)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # A non-integer follower count is impossible: these are European
        # thousands separators (e.g. "51.847" = 51,847) that Excel stored as
        # the float 51.847. Scale back to whole followers.
        if float(v) != int(v):
            return int(round(float(v) * 1000)), str(raw)
        return int(v), str(raw)
    s = str(raw)
    if not s:
        return None, ""
    t = s.lower().replace(",", "").replace("+", "").replace("members", "").strip()
    m = re.match(r"^([\d.]+)\s*([km])?$", t)
    if not m:
        return None, s
    num, suf = m.group(1), m.group(2)
    try:
        if suf == "k":
            return int(round(float(num) * 1_000)), s
        if suf == "m":
            return int(round(float(num) * 1_000_000)), s
        # No K/M suffix. A bare "51.847" or "105.371" is a European thousands
        # separator (LinkedIn sheet), NOT a decimal. One dot + exactly 3 trailing
        # digits => strip the dot. Otherwise treat as a plain number.
        if re.fullmatch(r"\d{1,3}\.\d{3}", num):
            return int(num.replace(".", "")), s
        return int(round(float(num))), s
    except ValueError:
        return None, s


def parse_price(v):
    """Return (low|None, high|None, on_request:bool, raw). A cell may list
    several deliverable prices ('Quote Tweet $7,000\\nThread $9,000'); we keep
    the min/max span plus the raw text so nothing is lost."""
    raw = clean(v)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v), int(v), False, str(int(v))
    s = str(raw)
    if s.upper().strip() in PRICE_ON_REQUEST:
        return None, None, True, s
    nums = [int(n.replace(",", "")) for n in re.findall(r"\$?\s*([\d,]{2,})", s)]
    nums = [n for n in nums if n >= 50]  # drop stray small ints (durations, etc.)
    if not nums:
        # non-empty but unparseable (e.g. 'Price on Request' variants)
        return None, None, True, s
    return min(nums), max(nums), False, s


def extract_handle(link):
    """Best-effort @handle from a social URL. Prefers X/Twitter; falls back to
    the handle segment of youtube/tiktok/instagram links. ARGUS keys on handles,
    so this is what lets a rate-card row bind to an audited subject."""
    s = clean(link)
    if not s or not isinstance(s, str):
        return None, None
    m = re.search(r"(?:twitter|x)\.com/@?([A-Za-z0-9_]{1,15})", s)
    if m:
        return "@" + m.group(1), "x"
    m = re.search(r"(?:tiktok\.com/@|instagram\.com/|youtube\.com/@)([A-Za-z0-9_.]+)", s)
    if m:
        plat = "tiktok" if "tiktok" in s else "instagram" if "instagram" in s else "youtube"
        return "@" + m.group(1).strip("/"), plat
    return None, None


def split_list(v):
    s = clean(v)
    if not s:
        return []
    parts = re.split(r"[,/]| and ", str(s))
    return [p.strip() for p in parts if p.strip()]


def price_range_level(v):
    s = clean(v)
    if isinstance(s, str) and set(s) == {"$"}:
        return len(s)
    return None


def to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def slugify(s):
    ascii_ = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_.lower()).strip("-")
    if slug:
        return slug
    # Non-ASCII-only name (e.g. Korean/CJK handles): a slug would be empty and
    # collapse every such KOL onto one id. Fall back to a stable hash so each
    # distinct name keeps a distinct, reproducible id.
    return "u-" + hashlib.sha1(str(s).encode("utf-8")).hexdigest()[:8]


def find_header_row(ws):
    for r in range(1, 8):
        cells = [clean(c) for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))]
        upper = {str(c).upper() for c in cells if c}
        if "KOL" in upper or "KOL NAME" in upper or ("NAME" in upper and "LINK" in upper):
            return r, cells
    return None, None


def pick_layout(header_upper):
    hset = set(header_upper)
    if "KOL NAME" in hset:
        return LAYOUT_LATAM
    if "NICHE" in hset and "KOL" not in hset:
        return LAYOUT_WEB2
    return LAYOUT_STANDARD


def ingest_sheet(ws):
    region = ws.title.strip().upper()
    hrow, header = find_header_row(ws)
    if not hrow:
        return []
    header_upper = [str(c).upper().strip() for c in header]
    layout = pick_layout(header_upper)
    col = {}  # field -> column index
    for i, h in enumerate(header_upper):
        if h in layout:
            col[layout[h]] = i

    out = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        name = clean(row[col["name"]]) if "name" in col and col["name"] < len(row) else ""
        if not name:
            continue

        def cell(field):
            i = col.get(field)
            return row[i] if i is not None and i < len(row) else None

        foll, foll_raw = parse_followers(cell("followers"))
        lo, hi, por, praw = parse_price(cell("pricing"))
        deliv = clean(cell("deliverables"))
        link = clean(cell("link")) or None
        handle, handle_platform = extract_handle(link)
        # Premium KOLs often leave PRICING as "-"/POR and list the actual rate
        # card inside DELIVERABLES ("Quote Tweet - $7,000 ..."). Recover that
        # spend signal rather than treating them as price-unknown.
        price_from_deliverables = False
        if lo is None and isinstance(deliv, str) and "$" in deliv:
            d_lo, d_hi, _, _ = parse_price(deliv)
            if d_lo is not None:
                lo, hi, por = d_lo, d_hi, False
                price_from_deliverables = True
        rec = {
            "id": f"{slugify(region)}--{slugify(name)}",
            "name": str(name),
            "region": region,
            "country": clean(cell("country")) or None,
            "platforms": split_list(cell("platforms")),
            "tier": to_int(cell("tier")),
            "followers": foll,
            "followers_raw": foll_raw or None,
            "categories": split_list(cell("categories")),
            "language": clean(cell("language")) or None,
            "rating": to_int(cell("rating")),
            "price_range": clean(cell("price_range")) or None,
            "price_range_level": price_range_level(cell("price_range")),
            "price_usd_low": lo,
            "price_usd_high": hi,
            "price_on_request": por,
            "price_from_deliverables": price_from_deliverables,
            "pricing_raw": praw or None,
            "deliverables": deliv or None,
            "link": link,
            "handle": handle,
            "handle_platform": handle_platform,
        }
        out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", help="path to the KOL .xlsx workbook")
    ap.add_argument("--replace", action="store_true",
                    help="drop all previously-ingested rows before adding these")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    fresh = []
    for ws in wb.worksheets:
        fresh.extend(ingest_sheet(ws))

    # Disambiguate genuine same-id rows within this workbook (e.g. the same
    # person listed once per platform) so no row is silently dropped.
    id_counts = {}
    for r in fresh:
        n = id_counts.get(r["id"], 0) + 1
        id_counts[r["id"]] = n
        if n > 1:
            r["id"] = f"{r['id']}--{n}"

    existing = {}
    meta_sources = []
    if OUT.exists() and not args.replace:
        prior = json.loads(OUT.read_text())
        for r in prior.get("records", []):
            existing[r["id"]] = r
        meta_sources = prior.get("sources", [])

    for r in fresh:
        existing[r["id"]] = r  # upsert

    src_name = Path(args.workbook).name
    meta_sources = [s for s in meta_sources if s.get("file") != src_name]
    meta_sources.append({
        "file": src_name,
        "ingested_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "rows": len(fresh),
    })

    records = sorted(existing.values(), key=lambda r: (r["region"], r["name"].lower()))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "count": len(records),
        "sources": meta_sources,
        "records": records,
    }, ensure_ascii=False, indent=2))

    # ---- summary to stderr for the operator ----
    regions = {}
    known_spend = 0
    priced = 0
    for r in records:
        regions[r["region"]] = regions.get(r["region"], 0) + 1
        if r["price_usd_low"]:
            priced += 1
            known_spend += r["price_usd_low"]
    print(f"wrote {OUT.relative_to(ROOT)}  ({len(records)} KOLs)", file=sys.stderr)
    print(f"  +{len(fresh)} rows from {src_name}", file=sys.stderr)
    print(f"  priced: {priced}/{len(records)}  (POR: {len(records)-priced})", file=sys.stderr)
    print(f"  floor cost to hire every priced KOL once: ${known_spend:,}", file=sys.stderr)
    print("  by region: " + ", ".join(f"{k} {v}" for k, v in sorted(regions.items())), file=sys.stderr)


if __name__ == "__main__":
    main()
